import requests
import pandas as pd
from datetime import datetime, timedelta, timezone
import json
import os
import schedule
import time
import threading
import re
import csv
from openpyxl.utils import get_column_letter
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.cell.cell import MergedCell


now = datetime.now()
year = now.year
month = now.month
day = now.day

# excel 데이터에서 심박, 호흡, 산포도의 평균을 구하는 함수
# 조건1) 평균을 구하는 값은 오전 8 ~ 오전 9시 데이터
# 조건2) 평균 계산 시 값이 0인 값들이 있으면 제외

# 현재는 scheduler를 통해 다운받은 데이터를 7~8, 15~ 16 시간대 평균 계산 
def get_average_data():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    data_dir = os.path.join(script_dir, "HospitalData")
    json_dir = os.path.join(script_dir, "json")

    now = datetime.now()
    today_str = now.strftime("%Y-%m-%d")

    target_times = [
        datetime.strptime(f"{today_str} 08:00", "%Y-%m-%d %H:%M"),
        datetime.strptime(f"{today_str} 16:00", "%Y-%m-%d %H:%M")
    ]

    # 병원별로 저장할 전체 결과 딕셔너리
    hospital_results = {}

    for filename in os.listdir(json_dir):
        if filename.endswith(".json"):
            hospital_name = filename.split('_')[0].upper()
            hospital_floor = filename.split('_')[3].split('.')[0]
            results = {
                "08:00": {"hr": [], "breath": [], "spo2": [], "rooms": []},
                "16:00": {"hr": [], "breath": [], "spo2": [], "rooms": []},
            }

            hospital_base_dir = os.path.join(data_dir, hospital_name) # HospitalData/YN

            if hospital_floor: #floor가 존재한다면 
                hospital_base_dir = os.path.join(data_dir, hospital_name, hospital_floor)

            today_date_dir = os.path.join(hospital_base_dir, today_str)

            if not os.path.isdir(today_date_dir):
                print(f"[{hospital_name}] 오늘 날짜 폴더 없음: {today_date_dir}")
                continue

            print(f"[{hospital_name}] 오늘 날짜 폴더 존재: {today_date_dir}")

            for data_file in os.listdir(today_date_dir):
                if not data_file.endswith(".csv"):
                    continue

                print(f"[{hospital_name}_{hospital_floor}] 처리 중: {data_file}")

                match = re.search(r"^(\d+_\d+)_", data_file)
                if match:
                    room_id = match.group(1)
                else:
                    print(f"[{hospital_name}_{hospital_floor}] 파일명에서 room_id 추출 실패: {data_file}")
                    continue

                file_path = os.path.join(today_date_dir, data_file)

                for target_time in target_times:
                    start_time = target_time - timedelta(hours=1)
                    time_key = target_time.strftime("%H:%M")

                    hr_list = []
                    breath_list = []
                    spo2_list = []

                    with open(file_path, 'r', encoding='utf-8') as csvfile:
                        reader = csv.reader(csvfile, delimiter=',')
                        next(reader, None)  # skip header

                        for i, row in enumerate(reader):
                            try:
                                timestamp_str = row[0].strip()
                                full_timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
                                timestamp = full_timestamp.replace(second=0)

                                if not (start_time <= timestamp <= target_time):
                                    continue

                                hr = int(row[3])
                                breath = int(row[4])
                                spo2 = int(row[5])

                                if hr != 0:
                                    hr_list.append(hr)
                                if breath != 0:
                                    breath_list.append(breath)
                                if spo2 != 0:
                                    spo2_list.append(spo2)

                            except (IndexError, ValueError) as e:
                                print(f"에러 ({hospital_name} {data_file} 줄 {i+1}): {e}")
                                continue

                    avg_hr = int(round(sum(hr_list) / len(hr_list))) if hr_list else 0
                    avg_breath = int(round(sum(breath_list) / len(breath_list))) if breath_list else 0
                    avg_spo2 = int(round(sum(spo2_list) / len(spo2_list))) if spo2_list else 0

                    results[time_key]["hr"].append(avg_hr)
                    results[time_key]["breath"].append(avg_breath)
                    results[time_key]["spo2"].append(avg_spo2)
                    results[time_key]["rooms"].append(room_id)

                    print(f"[{hospital_name}] [{time_key}] {room_id} → HR: {avg_hr}, BR: {avg_breath}, SPO2: {avg_spo2}")

            hospital_results[hospital_name] = {
                "floor": hospital_floor,
                "08:00": results["08:00"],
                "16:00": results["16:00"]
            }

    return hospital_results

def get_average_from_custom_folder():
    try:
        folder_path = input("csv를 저장한 폴더 경로를 입력하세요.")
        folder_path = os.path.abspath(folder_path)
        input_str = input("평균을 구할 시간을 입력하세요 (예: 2025-05-20 09:00): ")
        input_datetime = datetime.strptime(input_str, "%Y-%m-%d %H:%M")
    except ValueError:
        print("입력 형식이 잘못되었습니다. 예: 2025-05-20 09:00")
        return [], [], [], []

    start_time = input_datetime - timedelta(hours=1)

    all_avg_hr = []
    all_avg_breath = []
    all_avg_spo2 = []
    room_id_list = []

    if not os.path.isdir(folder_path):
        print(f"지정한 폴더가 존재하지 않습니다: {folder_path}")
        return [], [], [], []

    for data_file in os.listdir(folder_path):
        if data_file.endswith(".csv"):
            print(f"처리 중: {data_file}")
            hr_list = []
            breath_list = []
            spo2_list = []

            match = re.search(r"^(\d+_\d+)_", data_file)
            room_id = match.group(1) if match else "Unknown"

            file_path = os.path.join(folder_path, data_file)

            with open(file_path, 'r', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile, delimiter=',')
                next(reader, None)  # 헤더 스킵

                for i, row in enumerate(reader):
                    try:
                        timestamp_str = row[0].strip()
                        full_timestamp = datetime.strptime(timestamp_str, "%Y-%m-%d %H:%M:%S")
                        timestamp = full_timestamp.replace(second=0)

                        if not (start_time <= timestamp <= input_datetime):
                            continue

                        hr = int(row[3])
                        breath = int(row[4])
                        spo2 = int(row[5])

                        if hr != 0:
                            hr_list.append(hr)
                        if breath != 0:
                            breath_list.append(breath)
                        if spo2 != 0:
                            spo2_list.append(spo2)


                    except (IndexError, ValueError) as e:
                        print(f"에러 발생 (줄 {i+1}): {e}")
                        continue

            avg_hr = int(round(sum(hr_list) / len(hr_list))) if hr_list else 0
            avg_breath = int(round(sum(breath_list) / len(breath_list))) if breath_list else 0
            avg_spo2 = int(round(sum(spo2_list) / len(spo2_list))) if spo2_list else 0

            all_avg_hr.append(avg_hr)
            all_avg_breath.append(avg_breath)
            all_avg_spo2.append(avg_spo2)
            room_id_list.append(room_id)

            print(f"{data_file} ({room_id}) 평균 심박: {avg_hr}, 호흡: {avg_breath}, 산소포화도: {avg_spo2}")

    return all_avg_hr, all_avg_breath, all_avg_spo2, room_id_list

# 심박, 호흡, 산포도 평균 값들을 엑셀에 저장하는 함수

def make_averdata_to_excel(avr_heart, avr_breath, avr_spo2, room_id, filename="averdata.xlsx"):
    # 리스트 길이 체크
    if not (len(avr_heart) == len(avr_breath) == len(avr_spo2) == len(room_id)):
        raise ValueError("모든 리스트 길이가 같아야 합니다.")
    
    # DataFrame 생성
    df = pd.DataFrame({
        "room_id": room_id,
        "심박수": avr_heart,
        "호흡수": avr_breath,
        "산소포화도": avr_spo2 
    })

    
    
    # 엑셀 저장 (덮어쓰기)
    df.to_excel(filename, index=False)
    print(f"{filename} 파일에 저장 완료")

    # excel 스타일 만들기
def sanitize_sheet_title(title):
    # Excel에서 허용되지 않는 문자 제거
    for char in ['\\', '/', '*', '[', ']', ':', '?']:
        title = title.replace(char, '_')
    return title[:31]  # Excel 시트명은 31자 제한 있음

def sanitize_sheet_title(title):
    return title.replace("/", "_").replace("\\", "_")

def _get_value_by_room(results, time_key, metric, room_id):
    return results.get(time_key, {}).get(metric, {}).get(room_id)


def write_heart_rate_excel_by_sheet(
    hospital_results,
    filename="심박수_전체요약_확장.xlsx",
    titles=["심박수", "호흡수", "산소포화도"],
    date_str=None
):

    def find_next_start_column(ws, group_width=6):
        max_col = ws.max_column
        return 4 if max_col < 4 else max_col + 1

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )

    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    if date_str is None:
        date_str = datetime.now().strftime("%m/%d")

    metric_keys = ["hr", "breath", "spo2"]
    group_width = 6

    def sanitize_sheet_title(title):
        return re.sub(r'[\\/*?:\[\]]', '_', title)

    def _get_value_by_room(results, time_key, metric, room_id):
        """room_id에 맞는 값 찾기, 없으면 0 반환"""
        try:
            rooms = results[time_key]["rooms"]
            metric_list = results[time_key][metric]
            if room_id in rooms:
                idx = rooms.index(room_id)
                return metric_list[idx] if idx < len(metric_list) else 0
        except (KeyError, IndexError):
            pass
        return 0

    def get_total_columns(ws):
        return ws.max_column

    def merge_border(cell, add_border):
        old = cell.border or Border()

        def pick(side_old, side_new):
            return side_new if (side_new and side_new.style) else side_old

        cell.border = Border(
            left=pick(old.left, add_border.left),
            right=pick(old.right, add_border.right),
            top=pick(old.top, add_border.top),
            bottom=pick(old.bottom, add_border.bottom),
        )

    for hospital_name, results in hospital_results.items():
        floor = results.get("floor", "층미지정")
        sheet_title = sanitize_sheet_title(f"{hospital_name}_{floor}")
        if sheet_title not in wb.sheetnames:
            ws = wb.create_sheet(title=sheet_title)
            headers = ["병실", "침대", "환자명"]
            for col_idx, header in enumerate(headers, start=1):
                col_letter = get_column_letter(col_idx)
                ws.merge_cells(f"{col_letter}1:{col_letter}4")
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thick_border
                ws.column_dimensions[col_letter].width = 10
        else:
            ws = wb[sheet_title]

        data_start_col = find_next_start_column(ws, group_width)


        for i, title in enumerate(titles):
            offset = i * group_width + data_start_col

            ws.merge_cells(start_row=1, start_column=offset, end_row=1, end_column=offset + 5)
            cell = ws.cell(row=1, column=offset)
            cell.value = title
            cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thick_border

            ws.merge_cells(start_row=2, start_column=offset, end_row=2, end_column=offset + 5)
            cell = ws.cell(row=2, column=offset)
            cell.value = date_str
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thick_border

            ws.merge_cells(start_row=3, start_column=offset, end_row=3, end_column=offset + 1)
            ws.cell(row=3, column=offset).value = "간호사"
            ws.merge_cells(start_row=3, start_column=offset + 2, end_row=3, end_column=offset + 3)
            ws.cell(row=3, column=offset + 2).value = "장비"
            ws.merge_cells(start_row=3, start_column=offset + 4, end_row=3, end_column=offset + 5)
            ws.cell(row=3, column=offset + 4).value = "편차"

            for j in range(6):
                cell = ws.cell(row=3, column=offset + j)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thick_border

            time_headers = ["8시", "16시"] * 3
            for j, h in enumerate(time_headers):
                col = offset + j
                left_style = 'thick' if j == 0 else 'thin'
                right_style = 'thick' if j == 5 else 'thin'
                cell = ws.cell(row=4, column=col)
                cell.value = h
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    top=Side(style='thick'),
                    bottom=Side(style='thick'),
                    left=Side(style=left_style),
                    right=Side(style=right_style),
                )
                if j >= 4:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", patternType="solid")

        rooms_8 = results.get("08:00", {}).get("rooms", [])
        rooms_16 = results.get("16:00", {}).get("rooms", [])
        all_rooms = sorted(set(rooms_8 + rooms_16))

        room_data = []
        for room_id in all_rooms:
            match = re.match(r"(\d+)[\-_](\d+)", room_id)
            if match:
                room, bed = match.groups()
                bed = str(int(bed))
            else:
                room = room_id
                bed = "1"
            room_data.append((room, bed, room_id))

        row_start = 5
        room_rows = defaultdict(list)

        for idx, (room, bed, room_id) in enumerate(room_data):
            row = row_start + idx
            room_rows[room].append(row)

            border_right_style = Border(
                left=Side(style='thin'),
                right=Side(style='thick'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            border_left_style = Border(
                left=Side(style='thick'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for i, metric in enumerate(metric_keys):
                offset = i * group_width + data_start_col
                cols = [offset + j for j in range(6)]

                val_8 = _get_value_by_room(results, "08:00", metric, room_id)

                #print(f"[디버그] {room_id} - 08:00 {metric}: {val_8}")

                val_16 = _get_value_by_room(results, "16:00", metric, room_id)

                cell1 = ws.cell(row=row, column=cols[0])
                cell1.border = border_left_style
                cell1.alignment = Alignment(horizontal="center", vertical="center")

                cell2 = ws.cell(row=row, column=cols[1])
                cell2.border = border_right_style
                cell2.alignment = Alignment(horizontal="center", vertical="center")

                cell3 = ws.cell(row=row, column=cols[2], value=val_8)
                cell3.border = border_left_style
                cell3.alignment = Alignment(horizontal="center", vertical="center")

                cell4 = ws.cell(row=row, column=cols[3], value=val_16)
                cell4.border = border_right_style
                cell4.alignment = Alignment(horizontal="center", vertical="center")

                formula1 = f"={ws.cell(row=row, column=cols[0]).coordinate}-{ws.cell(row=row, column=cols[2]).coordinate}"
                formula2 = f"={ws.cell(row=row, column=cols[1]).coordinate}-{ws.cell(row=row, column=cols[3]).coordinate}"

                cell5 = ws.cell(row=row, column=cols[4], value=formula1)
                cell5.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", patternType="solid")
                cell5.border = border_left_style

                cell6 = ws.cell(row=row, column=cols[5], value=formula2)
                cell6.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", patternType="solid")
                cell6.border = border_right_style

            is_last_row = idx == len(room_data) - 1
            current_room_number = room.split('-')[0].split('_')[0]
            next_room_number = room_data[idx + 1][0].split('-')[0].split('_')[0] if not is_last_row else None
            use_thick = current_room_number != next_room_number
        
            total_cols = get_total_columns(ws)
            for col in range(1, total_cols + 1):
                cell = ws.cell(row=row, column=col)
                if not isinstance(cell, MergedCell):
                    # 병실, 침대, 환자명 열에만 값 설정
                    if col == 1:
                        cell.value = room
                    elif col == 2:
                        cell.value = bed
                    elif col == 3:
                        cell.value = ""  # 환자명은 빈칸
                    # 그 외 열은 값 덮어쓰지 않음

                    # 공통 스타일 적용
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    bottom_border = Border(bottom=Side(style='thick')) if use_thick else Border(bottom=Side(style='thin'))
                    merge_border(cell, bottom_border)

                    if col == 1:
                        merge_border(cell, Border(bottom=Side(style='thick'),left=Side(style='thick'), right=Side(style='thin')))
                    elif col == 2:
                        merge_border(cell, Border(left=Side(style='thick'), right=Side(style='thick')))

      
        # 병실 열 병합
        for room, rows in room_rows.items():
            if len(rows) > 1:
                ws.merge_cells(start_row=rows[0], start_column=1, end_row=rows[-1], end_column=1)
                    

    if wb.sheetnames:
        last_sheetname = wb.sheetnames[-1]
        last_ws = wb[last_sheetname]

        for col_idx in range(1, 4):
            for row in range(1, 5):
                cell = last_ws.cell(row=row, column=col_idx)
                cell.border = thick_border

        total_columns = last_ws.max_column
        for row in range(1, 5):
            for col in range(4, total_columns + 1):
                cell = last_ws.cell(row=row, column=col)
                cell.border = thick_border

    try:
        wb.save(filename)
        print(f"[완료] 엑셀 저장됨: {filename}")
    except PermissionError:
        print(f"[오류] 파일이 열려 있어 저장할 수 없습니다: {filename}")


def fetch_prometheus_metrics(job_name, metrics, start_time, end_time, step=10, timezone_offset=9, output_csv=None, room_id=None):
    """
    Parameters:
        job_name (str): Prometheus job 이름
        metrics (list): 조회할 metric 리스트
        start_time (datetime): 조회 시작 시간 (datetime 객체)
        end_time (datetime): 조회 종료 시간 (datetime 객체)
        step (int): step 간격 (초 단위)
        timezone_offset (int): 타임존 오프셋 (기본 KST: +9)
        output_csv (str): 저장할 CSV 디렉토리 경로 (None이면 저장하지 않음)
        room_id (str): CSV 파일명에 사용할 ID (예: "room1")

    Returns:
        pd.DataFrame 또는 None
    """
  

    query_url = ''
    kst = timezone(timedelta(hours=timezone_offset))

    start = int(start_time.timestamp())
    end = int(end_time.timestamp())

    dataframes = []

    for metric in metrics:
        query = f'{metric}{{job="{job_name}"}}'
        params = {
            "query": query,
            "start": start,
            "end": end,
            "step": step
        }

        response = requests.get(query_url, params=params)

        if response.status_code == 200:
            res_data = response.json()
            results = res_data["data"]["result"]

            for idx, series in enumerate(results):
                metric_name = series["metric"].get("__name__", metric)
                values = series["values"]

                df = pd.DataFrame(values, columns=["timestamp", metric_name])
                df["timestamp"] = pd.to_datetime(df["timestamp"], unit="s", utc=True).dt.tz_convert('Asia/Seoul')
                df["timestamp"] = df["timestamp"].dt.strftime('%Y-%m-%d %H:%M:%S')
                df[metric_name] = pd.to_numeric(df[metric_name], errors="coerce")
                df.set_index("timestamp", inplace=True)

                dataframes.append(df)
        else:
            print(f"[ERROR] {metric} 응답 실패: {response.status_code}")

    if dataframes:
        # 데이터프레임들을 시간 기준으로 병합
        merged_df = pd.concat(dataframes, axis=1).sort_index()

        # CSV 저장 (한 번만)
        if output_csv:
            # 년, 월, 일 받아오기
            uid = job_name.split('/')[1]
            os.makedirs(output_csv, exist_ok=True)
            csv_file_path = os.path.join(output_csv, f"{room_id}_{uid}.csv")
            merged_df.to_csv(csv_file_path)

        return merged_df
    else:
        print("데이터가 없습니다.")
        return None
    
#16시 10분 마다 데이터 저장 함수
def save_all_hospital_data():
    today_str = datetime.now().strftime("%Y-%m-%d")
    # 엑셀 데이터 저장 경로 설정
    script_dir = os.path.dirname(os.path.abspath(__file__))  # 현재 스크립트 파일의 경로
    uid_dir = os.path.join(script_dir, "json")  # JSON 파일들이 들어있는 폴더 경로

    today = datetime.now()
    year = today.year
    month = today.month
    day = today.day
    
    for filename in os.listdir(uid_dir):
        if filename.endswith(".json"):
            file_path = os.path.join(uid_dir, filename)

            # 병원명 추출 (예: 'yn_uid_list_2.json' → 'Yn')
            hospital_name = filename.split('_')[0].upper()  # 대문자로 통일: 'YN', 'HYO', 'JJ' 등
            hospital_floor = filename.split('_')[3].split('.')[0] # 층만 출력 (yn 일때만 사용)

            # 저장 경로 구성(기본 저장 경로 HospitalData/병원 이름/오늘 날짜 , floor존재시 추가)
            base_path = os.path.join(script_dir,"HospitalData",hospital_name)

            if hospital_floor:
                base_path = os.path.join(script_dir,"HospitalData",hospital_name,hospital_floor)

            base_path = os.path.join(script_dir,"HospitalData",hospital_name,hospital_floor,today_str)
            os.makedirs(base_path, exist_ok=True)

            # JSON 데이터 로드
            uid_data = load_json(file_path)

            for room_id, uid in uid_data.items():
                job_name1 = uid.split('/')[0]  # job name : '21b7'
                uid_value = uid.split('/')[1]  # uid : '0559E31031701'
                df = fetch_prometheus_metrics(
                    job_name= (job_name1 + '/' + uid_value ),
                    metrics=metrics,
                    start_time=yesterday_16pm,  
                    end_time=today_16pm,
                    step=10,
                    room_id=room_id
                )

                # 엑셀로 데이터 저장
                df.to_csv(os.path.join(base_path, f"{room_id}_{uid_value}.csv"))

        # scheduler를 통해 받아온 데이터 평균 계산
        results = get_average_data()
        # 엑셀 파일로 저장
        #make_scheduler_averdata_to_excel_by_hospital(results)
        write_heart_rate_excel_by_sheet(results)

# 입력 시간으로 데이터를 추출 및 저장하는 함수
def save_input_hospital_data():
    hospital_input = input("병원 코드를 입력하세요 (ex: hyo, jj, yn5): ").strip()

    hospital_code = ''.join(filter(str.isalpha, hospital_input))
    hospital_num = ''.join(filter(str.isdigit, hospital_input))

    if hospital_code == 'yn' and hospital_num:
        expected_filename = f"{hospital_code}_uid_list_{hospital_num}.json"
    else:
        expected_filename = f"{hospital_code}_uid_list.json"

    # 실행 중인 스크립트의 절대 경로 기준으로 경로 설정
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(script_dir, 'json', expected_filename)

    print(f"filepath : {file_path}")

    if not os.path.exists(file_path):
        print(f"파일을 찾을 수 없습니다: {file_path}")
        return

    input_str = input("시간을 입력하세요 (예: 20250520 0900): ")
    input_datetime = datetime.strptime(input_str, "%Y%m%d %H%M")
    yesterday_datetime = input_datetime - timedelta(days=1)
    
    output_csv1 = input("엑셀 저장 경로를 입력하세요 :")

    uid_data = load_json(file_path)

    for room_id, uid in uid_data.items():
        job_name = uid.split('/')[0]  # 예: '21b7'
        uid_value = uid.split('/')[1]  # 예: '0559E31031701'
        job_name1 = f"{job_name}/{uid_value}"
    
        df = fetch_prometheus_metrics(
            job_name=job_name1,
            metrics=metrics,
            start_time=yesterday_datetime,
            end_time=input_datetime,
            step=10,
            output_csv=output_csv1,
            room_id=room_id,
        )
   
# run_scheduler 함수
def run_scheduler():
    while True:
        schedule.run_pending()
        time.sleep(1)   

# json 파일 열기
def load_json(file_path):
    if not os.path.exists(file_path):
        print(f"[INFO] '{file_path}' 파일이 없어 새로 만듭니다.")
        return {}
    with open(file_path, 'r', encoding='utf-8') as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            print(f"[WARNING] '{file_path}'이 비어있거나 손상되어 초기화합니다.")
            return {}
        
# json 파일 저장
def save_json(file_path, data):
    try:
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        print(f"[INFO] JSON saved successfully to '{file_path}'.")
    except Exception as e:
        print(f"[ERROR] Failed to save JSON to '{file_path}': {e}")

#json 파일 추가 및 수정
def add_or_change_uid(file_path,room_id,uid):
    data = load_json(file_path)
    data[room_id] = uid
    save_json(file_path, data)
    print(f"[INFO] UID '{uid}' has been added/updated.")

#json 파일 삭제
def delete_json(file_path, room_id):
    data = load_json(file_path)

    if room_id not in data:
        print(f"[WARN] UID '{room_id}' not found.")
        return

    print("삭제 옵션을 선택하세요:")
    print("1. UID만 삭제 (ROOM_ID는 유지)")
    print("2. ROOM_ID와 UID 모두 삭제")

    choice = input("입력 (1 또는 2): ").strip()

    if choice == "1":
        data[room_id] = ""
        save_json(file_path, data)
        print(f"[INFO] UID '{room_id}'의 value만 삭제되었습니다.")
    elif choice == "2":
        del data[room_id]
        save_json(file_path, data)
        print(f"[INFO] UID '{room_id}' key와 value 모두 삭제되었습니다.")
    else:
        print("[ERROR] 잘못된 입력입니다. 삭제가 취소되었습니다.")


# -------------------------
# 실행 구문
# -------------------------
if __name__ == "__main__":
    # 시간 설정 (KST 기준 어제 16시 ~ 오늘 16시)
    kst = timezone(timedelta(hours=9))
    today_16pm = datetime.now(tz=kst).replace(hour=16, minute=0, second=0, microsecond=0)
    yesterday_16pm = today_16pm - timedelta(days=1)
    lhour_ago = today_16pm -timedelta(hours=1)
    # metric 및 job 설정
    metrics = [
        'radar_v3_state', 'radar_v3_heart_detection', 'radar_v3_heart',
        'radar_v3_breath', 'radar_v3_sp02', 'radar_v3_drop', 'radar_v3_radar_rssi'
    ]
   
    # At 16:10 save all hospital data
    # Keep the main thread alive so the scheduler can run
    schedule.every().day.at("12:45").do(save_all_hospital_data)
    thread = threading.Thread(target=run_scheduler)
    thread.daemon = True
    thread.start()

    aver_hr=[]
    aver_br = []
    aver_spo2 = []
    room_id_list = []
# command를 받아서 데이터를 저장 및 평균 계산


    
def get_full_json_path():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(script_dir, 'json')

    filename = input("파일 이름을 입력하세요 (예: yn_uid_list_2.json): ").strip()
    filename = os.path.basename(filename)  # ⚠️ 여기서 디렉토리 제거

    filepath = os.path.join(file_path, filename)
    print(f"[INFO] JSON 파일 경로: {filepath}")
    return filepath

while True:
    print("\n명령어를 입력하세요:")
    print(" - command : 병원 데이터 다운로드")
    print(" - average : 호실별 평균 계산 및 엑셀 저장")
    print(" - add     : UID 추가 또는 변경")
    print(" - delete  : UID 삭제")
    print(" - show    : UID 목록 보기")
    print(" - exit    : 프로그램 종료")
    command = input(">>> ").strip().lower()

    if command == "command":
        save_input_hospital_data()

    elif command == "average":
        aver_hr, aver_br, aver_spo2, room_id_list = get_average_from_custom_folder()
        make_averdata_to_excel(aver_hr, aver_br, aver_spo2, room_id_list)

    elif command == "add":
        file_path = get_full_json_path()
        room_id = input("추가할 병실 ID를 입력하세요 (예: 211_1): ").strip()
        uid = input("UID 값을 입력하세요 (예: 21b7/A1B2C3D4): ").strip()
        add_or_change_uid(file_path, room_id, uid)

    elif command == "delete":
        file_path = get_full_json_path()
        room_id = input("삭제할 병실 ID를 입력하세요 (예: 211_1): ").strip()
        delete_json(file_path, room_id)

    elif command == "show":
        file_path = get_full_json_path()
        data = load_json(file_path)
        if not data:
            print("[INFO] UID 데이터가 비어 있습니다.")
        else:
            print("[UID 목록]")
            for room_id, uid in data.items():
                print(f" - {room_id}: {uid}")

    elif command == "exit":
        print("프로그램 종료 중...")
        break

    else:
        print("알 수 없는 명령어입니다.")

print("메인 루프 종료, 프로그램 종료.")
